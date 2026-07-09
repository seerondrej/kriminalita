#!/usr/bin/env python3
"""
Parse Police ČR yearly crime files (2008-2025) into one unified JSON.

Two eras with different Excel layouts:
  * Legacy 2008-2015 (.xls): sheet 0 = Česká republika, coded region sheets
      col1 = TSK code (float), col2 = name, col3 = Zjištěno (registered),
      col6 = Objasněno (solved). Category subtotals coded like "101-190".
  * Modern 2016-2025 (.xlsx): sheet "Česká republika" + named kraj sheets
      col0 = TSK code, col1 = name, col2 = REGISTROVÁNO, col3 = OBJASNĚNO.
      Category subtotals coded like "100-199".

Output: data/crime-data.json  (national + regional, by category, total and per TSK).
"""
import json, os, re
import openpyxl, xlrd

RAW = os.path.join(os.path.dirname(__file__), "..", "data", "raw")
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "crime-data.json")

# Canonical categories. Each era encodes the subtotal under a different range
# code, and some modern years drift (e.g. total is 0-999 or 100-999) -> use sets.
CATEGORIES = [
    # key, label, code set (all eras), tsk lo, tsk hi, color
    ("nasilna",     "Násilná kriminalita",     {"101-190", "100-199"}, 100, 199, "#e5484d"),
    ("mravnostni",  "Mravnostní kriminalita",  {"201-290", "200-299"}, 200, 299, "#d6409f"),
    ("majetkova",   "Majetková kriminalita",   {"311-590", "300-599"}, 300, 599, "#f5a623"),
    ("ostatni",     "Ostatní kriminalita",     {"611-690", "611-664", "600-699"}, 600, 699, "#8e4ec6"),
    ("zbyvajici",   "Zbývající kriminalita",   {"721-790", "700-799"}, 700, 799, "#0091ff"),
    ("hospodarska", "Hospodářská kriminalita", {"801-890", "800-899"}, 800, 899, "#12a594"),
]
OBECNA = ("obecna", "Obecná kriminalita", {"101-690", "101-664", "100-699"})
CELKOVA = ("celkova", "Celková kriminalita", {"101-903", "101-902", "0-999", "100-999"})

# Kraj canonical code + name. Region sheets are matched by a keyword found in the
# sheet's label cell (legacy) or its sheet name (modern) -> era-independent.
KRAJE = [
    # code, name, uppercase keyword to detect
    ("PHA", "Praha",           "PRAH"),
    ("STC", "Středočeský",     "STŘEDOČESK"),
    ("JHC", "Jihočeský",       "JIHOČESK"),
    ("PLK", "Plzeňský",        "PLZE"),
    ("KVK", "Karlovarský",     "KARLOVARSK"),
    ("ULK", "Ústecký",         "ÚSTECK"),
    ("LBK", "Liberecký",       "LIBERECK"),
    ("HKK", "Královéhradecký", "KRÁLOVÉHRADECK"),
    ("PAK", "Pardubický",      "PARDUBICK"),
    ("VYS", "Vysočina",        "VYSOČIN"),
    ("JHM", "Jihomoravský",    "JIHOMORAVSK"),
    ("OLK", "Olomoucký",       "OLOMOUCK"),
    ("ZLK", "Zlínský",         "ZLÍNSK"),
    ("MSK", "Moravskoslezský", "MORAVSKOSLEZSK"),
]

def match_region(label):
    """Map a sheet label / name to a canonical region code, or 'CZ', or None."""
    if not label:
        return None
    up = str(label).upper()
    if "ČESKÁ REPUBLIKA" in up:
        return "CZ"
    for code, _name, kw in KRAJE:
        if kw in up:
            return code
    return None

def cat_of_tsk(code):
    for key, _l, _codes, lo, hi, _c in CATEGORIES:
        if lo <= code <= hi:
            return key
    return None

def norm_code(v):
    """Return ('int', n) for a plain TSK code, ('range', 'a-b') for subtotal, or None."""
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    if re.fullmatch(r"\d+-\d+", s):
        return ("range", s)
    # floats like '101.0' or ints
    m = re.fullmatch(r"(\d+)(\.0+)?", s)
    if m:
        return ("int", int(m.group(1)))
    return None

def num(v):
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return None

# ---- per-sheet row extractor, returns dicts keyed by code ----
def extract_modern(rows):
    """rows: list of tuples. col0 code, col1 name, col2 reg, col3 solved."""
    tsk, ranges = {}, {}
    for r in rows:
        if len(r) < 3:
            continue
        nc = norm_code(r[0])
        if not nc:
            continue
        name = str(r[1]).replace("\n", " ").strip() if r[1] is not None else ""
        reg = num(r[2]); sol = num(r[3]) if len(r) > 3 else None
        if nc[0] == "int":
            tsk[nc[1]] = {"name": name, "reg": reg, "sol": sol}
        else:
            ranges[nc[1]] = {"reg": reg, "sol": sol}
    return tsk, ranges

def extract_legacy(sheet, sol_col):
    """col1 code, col2 name, col3 Zjištěno (registered); solved column era-specific
    (vusc 2008-09 -> col4, a_I 2010-15 -> col6)."""
    tsk, ranges = {}, {}
    for i in range(sheet.nrows):
        nc = norm_code(sheet.cell_value(i, 1))
        if not nc:
            continue
        name = str(sheet.cell_value(i, 2)).replace("\n", " ").strip()
        reg = num(sheet.cell_value(i, 3))
        sol = num(sheet.cell_value(i, sol_col)) if sheet.ncols > sol_col else None
        if nc[0] == "int":
            tsk[nc[1]] = {"name": name, "reg": reg, "sol": sol}
        else:
            ranges[nc[1]] = {"reg": reg, "sol": sol}
    return tsk, ranges

def load_year(year):
    xlsx = os.path.join(RAW, f"{year}.xlsx")
    xls = os.path.join(RAW, f"{year}.xls")
    sheets = {}  # canonical region code -> (tsk, ranges); 'CZ' for national
    if os.path.exists(xlsx):
        wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
        for sn in wb.sheetnames:
            key = match_region(sn)
            if not key:
                continue
            rows = list(wb[sn].iter_rows(values_only=True))
            sheets[key] = extract_modern(rows)
        era = "modern"
    elif os.path.exists(xls):
        wb = xlrd.open_workbook(xls)
        # 2008-2009 use vusc_* sheets with Objasněno in col4; 2010-2015 use a_I_* with col6
        sol_col = 4 if any(s.startswith("vusc") for s in wb.sheet_names()) else 6
        for sn in wb.sheet_names():
            s = wb.sheet_by_name(sn)
            label = str(s.cell_value(4, 2)).strip() if s.nrows > 4 else ""
            key = match_region(label)
            if not key:
                continue
            sheets[key] = extract_legacy(s, sol_col)
        era = "legacy"
    else:
        return None, None
    return sheets, era

def build():
    years = list(range(2008, 2026))
    data = {
        "national": {"total": {}, "byCategory": {}, "byTsk": {}},
        "regional": {},
        "regionsMeta": [{"code": c, "name": n} for c, n, _kw in KRAJE],
        "categoriesMeta": (
            [{"key": k, "label": l, "color": col} for k, l, _codes, _lo, _hi, col in CATEGORIES]
            + [{"key": OBECNA[0], "label": OBECNA[1], "color": "#64748b"}]
        ),
        "years": years,
        "eraBreak": 2016,
    }
    for c, n, _kw in KRAJE:
        data["regional"][c] = {"total": {}, "byCategory": {}, "byTsk": {}}
    tsk_names = {}

    for year in years:
        sheets, era = load_year(year)
        if not sheets:
            print(f"  {year}: MISSING")
            continue
        cnt_tsk = 0
        def pick(ranges, codeset):
            for c in codeset:
                if c in ranges:
                    return ranges[c]
            return None

        for region, (tsk, ranges) in sheets.items():
            target = data["national"] if region == "CZ" else data["regional"][region]
            # total
            tot = pick(ranges, CELKOVA[2])
            if tot:
                target["total"][str(year)] = tot
            # categories (6 + obecna)
            for key, _lab, codes, _lo, _hi, _col in CATEGORIES:
                v = pick(ranges, codes)
                if v:
                    target["byCategory"].setdefault(key, {})[str(year)] = v
            ov = pick(ranges, OBECNA[2])
            if ov:
                target["byCategory"].setdefault(OBECNA[0], {})[str(year)] = ov
            # individual TSK codes
            for code, vals in tsk.items():
                cat = cat_of_tsk(code)
                entry = {"reg": vals["reg"], "sol": vals["sol"]}
                target["byTsk"].setdefault(str(code), {})[str(year)] = entry
                if region == "CZ" and vals["name"] and code not in tsk_names:
                    tsk_names[code] = {"name": vals["name"], "cat": cat}
                cnt_tsk += 1
        print(f"  {year} [{era}]: regions={len(sheets)} tsk-cells={cnt_tsk} "
              f"total={data['national']['total'].get(str(year))}")

    # attach TSK metadata (national names) sorted by code
    data["tskMeta"] = [
        {"code": c, "name": tsk_names[c]["name"], "cat": tsk_names[c]["cat"]}
        for c in sorted(tsk_names)
    ]
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    kb = os.path.getsize(OUT) / 1024
    print(f"\nWrote {OUT}  ({kb:.0f} KB)")
    print(f"TSK codes: {len(data['tskMeta'])}, years with national total: "
          f"{len(data['national']['total'])}")

if __name__ == "__main__":
    print("Parsing Police ČR yearly files...")
    build()
